"""
Excel 檔案解析功能
"""
import os
import time
import zipfile
import xml.etree.ElementTree as ET
import re
import json
import hashlib
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
import config.settings as settings
from utils.cache import copy_to_cache
import logging
import urllib.parse

def extract_external_refs(xlsx_path, disallow_mainproc_xml_fallback: bool = True):
    """
    解析 Excel xlsx 中 external reference mapping: [n] -> 路徑
    支援兩種來源：
    - xl/externalLinks/externalLinkN.xml 的 externalBookPr@href
    - xl/externalLinks/_rels/externalLinkN.xml.rels 中 Relationship@Target
    
    優先使用子進程隔離，失敗時回退到直接解析
    """
    # 嘗試使用子進程隔離
    try:
        from utils.xml_subproc import is_xml_subprocess_enabled, extract_external_refs_subprocess
        if is_xml_subprocess_enabled():
            if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                print(f"[excel_parser] extract_external_refs using subprocess for {os.path.basename(xlsx_path)}")
            return extract_external_refs_subprocess(xlsx_path)
    except Exception as e:
        if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
            print(f"[excel_parser] subprocess failed, fallback to direct parsing: {e}")
    
    # 若不允許主進程 XML 回退，直接返回空映射
    if disallow_mainproc_xml_fallback:
        if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
            print(f"[excel_parser] extract_external_refs direct parsing disabled; returning empty map")
        return {}
    
    # 回退到直接解析（允許時）
    if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
        print(f"[excel_parser] extract_external_refs using direct parsing for {os.path.basename(xlsx_path)}")
    
    ref_map = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            for rel in rels.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.attrib.get('Type','').endswith('/externalLink'):
                    target = rel.attrib.get('Target','')  # e.g., externalLinks/externalLink1.xml
                    m = re.search(r'externalLink(\d+)\.xml', target)
                    if not m:
                        continue
                    num = int(m.group(1))
                    path = ''
                    # 1) 嘗試 externalLinkN.xml 的 externalBookPr@href
                    try:
                        link_xml = z.read(f'xl/{target}')
                        link_tree = ET.fromstring(link_xml)
                        book_elem = link_tree.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}externalBookPr')
                        if book_elem is not None:
                            path = book_elem.attrib.get('href', '')
                    except Exception:
                        pass
                    # 2) 若仍無，嘗試 externalLinks/_rels/externalLinkN.xml.rels 的 Relationship@Target
                    if not path:
                        try:
                            rels_path = f"xl/externalLinks/_rels/externalLink{num}.xml.rels"
                            if rels_path in z.namelist():
                                link_rels_xml = z.read(rels_path)
                                link_rels = ET.fromstring(link_rels_xml)
                                rel_node = link_rels.find('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
                                if rel_node is not None:
                                    path = rel_node.attrib.get('Target','')
                        except Exception:
                            pass
                    ref_map[num] = path or ''
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as e:
        logging.error(f"提取外部參照時發生錯誤: {xlsx_path}, 錯誤: {e}")
    return ref_map

def _normalize_path(p: str) -> str:
    if not p:
        return p
    s = urllib.parse.unquote(p.strip())
    # Handle file: scheme robustly
    try:
        u = urllib.parse.urlparse(s)
        if u.scheme == 'file':
            if u.netloc:  # UNC: file://server/share/path
                path_part = u.path.lstrip('/').replace('/', '\\')
                s = "\\\\" + u.netloc + "\\" + path_part
            else:  # local: file:///C:/path or file:/C:/path or file:\C:\path
                rest = u.path or s[5:]
                rest = rest.lstrip('/\\')
                s = rest.replace('/', '\\')
    except Exception:
        pass
    # Fallback: strip 'file:' prefix crudely if present
    if s.lower().startswith('file:'):
        s = s[5:].lstrip('/\\')
    # normalize backslashes
    s = s.replace('/', '\\')
    # collapse duplicate backslashes but keep UNC prefix
    if s.startswith('\\\\'):
        prefix = '\\'
        t = s[2:]
        while '\\' in t:
            t = t.replace('\\\\', '\\')
        s = '\\' + t
    else:
        while '\\' in s and '\\\\' in s:
            s = s.replace('\\\\', '\\')
    return s


def _excel_external_prefix(norm_path: str, sheet: str) -> str:
    """
    將歸一化路徑與工作表組裝為 Excel 標準外部參照前綴：
    'C:\\dir\\[Workbook.xlsx]Sheet Name'
    注意：整段（目錄 + [檔名] + 工作表）以單引號包裹；工作表名中的單引號需轉義為兩個單引號。
    """
    if not norm_path:
        return None
    # 分割目錄與檔名
    base = os.path.basename(norm_path)
    dir_ = os.path.dirname(norm_path)
    # 若 base 沒有副檔名，原樣處理
    fname = base
    sheet_escaped = (sheet or '').replace("'", "''")
    inside = ''
    if dir_:
        inside = dir_.rstrip('\\') + '\\'
    inside += f"[{fname}]" + sheet_escaped
    return f"'{inside}'"


def pretty_formula(formula, ref_map=None):
    """
    將公式中的外部參照 [n]Sheet! 還原為 'full\\normalized\\path'!Sheet! 的可讀形式。
    同時保留 Excel 語法結構，避免造成假差異。
    """
    if formula is None:
        return None
    
    # 修改：處理 ArrayFormula 物件
    if isinstance(formula, ArrayFormula):
        formula_str = formula.text if hasattr(formula, 'text') else str(formula)
    else:
        formula_str = str(formula)
    
    if ref_map:
        # 1) 直接替換形如 [n]Sheet! 為 'path'!Sheet!
        def repl_path_with_sheet(m):
            n = int(m.group(1))
            sheet = m.group(2)
            # 清理工作表名左右可能存在的引號，避免重複引號（例如 'Sheet 2' 內層再包一層）
            try:
                if sheet is not None:
                    sheet = str(sheet).strip().strip("'\"")
            except Exception:
                pass
            raw_path = ref_map.get(n, '')
            norm_path = _normalize_path(raw_path)
            if norm_path:
                prefix = _excel_external_prefix(norm_path, sheet)
                # 保證只有一個單引號包裹在 ! 之前：處理邊界 "''!" → "'!"
                out = f"{prefix}!"
                out = out.replace("''!", "'!")
                return out
            return m.group(0)
        s = re.sub(r"\[(\d+)\]([^!\]]+)!", repl_path_with_sheet, formula_str)
        
        # 2) 對其餘殘留的 [n] 標記（未帶 sheet 名）插入可讀提示
        def repl_annotate(m):
            n = int(m.group(1))
            raw_path = ref_map.get(n, '')
            norm_path = _normalize_path(raw_path)
            if norm_path:
                return f"[外部檔案{n}: {norm_path}]"
            return m.group(0)
        s = re.sub(r"\[(\d+)\]", repl_annotate, s)
        # 邊界清理：避免在等號後緊接本機/UNC 路徑時出現兩個單引號（="'"'C:\... → ='C:\...）
        try:
            s = re.sub(r"=\s*''(?=(?:[A-Za-z]:\\|\\\\))", "='", s)
        except Exception:
            pass
        return s
    else:
        return formula_str

def get_cell_formula(cell):
    """
    取得 cell 公式（不論係普通 formula or array formula），一律回傳公式字串
    """
    if cell.data_type == 'f':
        if isinstance(cell.value, ArrayFormula):
            # 修改：返回 ArrayFormula 的實際公式字符串，而不是物件
            return cell.value.text if hasattr(cell.value, 'text') else str(cell.value)
        return cell.value
    return None

def serialize_cell_value(value):
    """
    序列化儲存格值
    """
    if value is None: 
        return None
    if isinstance(value, ArrayFormula): 
        return None
    if isinstance(value, datetime): 
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)): 
        return value
    return str(value)

def get_excel_last_author(path):
    """
    以非鎖定方式讀取 Excel 檔案的最後修改者（安全優先）：
    - 預設優先使用子進程讀取 meta（避免主進程 ET/openpyxl 風險）。
    - 如啟用嚴格/僅子進程模式，失敗則直接返回 None，不在主進程做 ET。
    - 如允許主進程 ET，才會在 GC guard 內做最小範圍的 zip+ET 解析。
    """
    try:
        if not getattr(settings, 'ENABLE_LAST_AUTHOR_LOOKUP', True):
            return None

        # 先複製到本地快取，避免直接打開原始檔案
        local_path = copy_to_cache(path, silent=True)
        if not local_path or not os.path.exists(local_path):
            return None

        # 1) 優先：子進程讀取 meta
        try:
            from utils.xml_subproc import is_xml_subprocess_enabled, read_meta_subprocess
            if is_xml_subprocess_enabled():
                meta = read_meta_subprocess(local_path)
                if isinstance(meta, dict):
                    la = (meta or {}).get('last_author')
                    if la:
                        return la
        except Exception:
            pass

        # 2) 嚴格安全模式：只允許子進程，失敗則返回 None
        if getattr(settings, 'LAST_AUTHOR_SUBPROCESS_ONLY', True) or getattr(settings, 'DISALLOW_MAINPROC_META_ET', True):
            return None

        # 3) 允許主進程 ET：在 GC guard 內做最小解析
        try:
            from utils.gc_guard import gc_guard_any_thread
            use_gc_guard = bool(getattr(settings, 'ENABLE_XML_GC_GUARD', True))
            with gc_guard_any_thread(enabled=use_gc_guard, do_collect=False):
                with zipfile.ZipFile(local_path, 'r') as z:
                    core_xml = z.read('docProps/core.xml')
                    root = ET.fromstring(core_xml)
                    ns = {
                        'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
                        'dc': 'http://purl.org/dc/elements/1.1/'
                    }
                    node = root.find('cp:lastModifiedBy', ns)
                    if node is None:
                        node = root.find('dc:lastModifiedBy', ns)
                    author = (node.text or '').strip() if node is not None else None
                    return author or None
        except Exception:
            pass

        return None

    except FileNotFoundError:
        logging.warning(f"檔案未找到: {path}")
        return None
    except PermissionError:
        logging.error(f"權限不足: {path}")
        return None
    except OSError as e:
        logging.error(f"Excel 檔案讀取 I/O 錯誤: {path}, {e}")
        return None

def safe_load_workbook(path, max_retry=3, delay=0.5, **kwargs):
    """
    安全載入 Excel 檔案，帶重試機制和錯誤診斷
    """
    last_err = None
    for i in range(max_retry):
        try:
            # 檢查文件是否可讀
            if not os.path.exists(path):
                raise FileNotFoundError(f"檔案不存在: {path}")
            
            if not os.access(path, os.R_OK):
                raise PermissionError(f"檔案無讀取權限: {path}")
            
            # 檢查文件大小
            file_size = os.path.getsize(path)
            if file_size == 0:
                raise ValueError(f"檔案為空: {path}")
            
            # 嘗試載入
            from utils.gc_guard import gc_guard_any_thread
            use_gc_guard = bool(getattr(settings, 'ENABLE_XML_GC_GUARD', True))
            with gc_guard_any_thread(enabled=use_gc_guard, do_collect=False):
                wb = load_workbook(path, **kwargs)
            return wb
            
        except PermissionError as e:
            last_err = e
            logging.warning(f"權限錯誤，重試 {i+1}/{max_retry}: {e}")
            time.sleep(delay)
            
        except (zipfile.BadZipFile, zipfile.LargeZipFile) as e:
            last_err = e
            logging.error(f"ZIP 檔案損壞: {path}, 錯誤: {e}")
            break
            
        except ET.ParseError as e:
            last_err = e
            logging.error(f"XML 解析錯誤: {path}, 錯誤: {e}")
            break
            
        except Exception as e:
            last_err = e
            error_type = type(e).__name__
            logging.error(f"載入 Excel 檔案時發生 {error_type} 錯誤: {path}, 錯誤: {e}")
            if i < max_retry - 1:
                time.sleep(delay)
            else:
                break
    
    raise last_err

def dump_excel_cells_with_timeout(path, show_sheet_detail=True, silent=False):  # noqa: C901
    """
    提取 Excel 檔案中的所有儲存格數據（含公式）
    - 會先將來源檔複製到本地快取，再以 openpyxl 讀取（絕不直接讀原檔，視設定而定）
    - 值引擎優先用 polars（如不可用則自動回退到 XML）
    - 修正：external_ref 先安全初始化為 False，避免 UnboundLocalError
    """
    # 更新全局變數
    settings.current_processing_file = path
    settings.processing_start_time = time.time()

    # 子進程完整掃描（no fallback）
    try:
        if not silent:
            print("   🔒 子進程完整掃描（no fallback）…")
        # 只處理快取副本
        local_path = copy_to_cache(path, silent=silent)
        if not local_path or not os.path.exists(local_path):
            if not silent:
                print("   ❌ 無法使用快取副本（子進程模式）")
            return None
        from utils.subprocess_manager import get_subprocess_manager
        mgr = get_subprocess_manager()
        include_vals = not bool(getattr(settings, 'FORMULA_ONLY_MODE', False))
        batch_sz = int(getattr(settings, 'EXCEL_BATCH_SIZE', 10000) or 10000)
        result = mgr.scan_excel_complete(local_path, include_formulas=True, include_values=include_vals, batch_size=batch_sz)
        if result and result.get('success'):
            excel_data = result.get('excel_data', {}) or {}
            # --- prettify external references using ref_map from subprocess (safe) ---
            try:
                try:
                    ref_map = extract_external_refs(local_path)
                except Exception:
                    ref_map = {}
                import re as _re
                _preview = 0
                for _ws, _wsdata in (excel_data or {}).items():
                    if not isinstance(_wsdata, dict):
                        continue
                    for _addr, _cell in _wsdata.items():
                        if not isinstance(_cell, dict):
                            continue
                        f0 = _cell.get('formula')
                        if f0 is not None:
                            s_before = None
                            try:
                                s_before = str(f0)
                                f_pretty = pretty_formula(f0, ref_map=ref_map)
                                _cell['formula'] = f_pretty
                            except Exception:
                                pass
                            try:
                                s_after = str(_cell.get('formula'))
                                ext = False
                                if s_before and _re.search(r"\[(\d+)\][^!\]]+!", s_before):
                                    ext = True
                                elif isinstance(s_after, str) and _re.search(r"'[^']*\\\[[^\\\]]+\][^']*'!", s_after):
                                    ext = True
                                elif isinstance(s_after, str) and _re.search(r"\[[^\]]+\][^!]+!", s_after):
                                    ext = True
                                _cell['external_ref'] = bool(ext)
                            except Exception:
                                pass
                            if getattr(settings, 'SHOW_DEBUG_MESSAGES', False) and _preview < 3:
                                try:
                                    print(f"   [prettify] {_ws}!{_addr}: {s_before} -> {_cell.get('formula')}")
                                except Exception:
                                    pass
                                _preview += 1
            except Exception:
                pass
            if not silent:
                try:
                    total_cells = sum(len(ws) for ws in excel_data.values())
                    total_ws = len(excel_data)
                    print(f"   ✅ 子進程掃描完成：worksheets={total_ws} cells={total_cells}")
                except Exception:
                    pass
            return excel_data
        else:
            if not silent:
                err = (result or {}).get('error', 'unknown')
                print(f"   ❌ 子進程掃描失敗：{err}（no fallback）")
            return None
    except Exception as _sube:
        if not silent:
            print(f"   ❌ 子進程掃描異常：{_sube}（no fallback）")
        return None

    wb = None
    try:
        if not silent:
            try:
                print(f"   📊 檔案大小: {os.path.getsize(path)/(1024*1024):.1f} MB")
            except Exception:
                pass

        # 只處理快取副本
        local_path = copy_to_cache(path, silent=silent)
        if not local_path or not os.path.exists(local_path):
            if not silent:
                print("   ❌ 無法使用快取副本（嚴格模式下不會讀取原檔），略過此檔案。")
            return None

        read_only_mode = True
        if not silent:
            print(f"   🚀 讀取模式: read_only={read_only_mode}, data_only=False")

        if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
            try:
                _sz = os.path.getsize(local_path)
                _kb = f"{_sz/1024:.1f} KB"
                _mb = f"{_sz/1024/1024:.2f} MB"
                _sz_str = f"bytes={_sz} ({_kb}, {_mb})"
            except Exception:
                _sz_str = 'bytes=N/A'
            print(f"   [xlsx-open] local_path={local_path} size={_sz_str}")
        # 在 FORMULA_ONLY_MODE 下先行子進程 openpyxl_scan，成功則避免主進程載入 workbook
        try:
            if bool(getattr(settings, 'FORMULA_ONLY_MODE', False)):
                from utils.xml_subproc import is_xml_subprocess_enabled, read_values_subprocess
                if is_xml_subprocess_enabled():
                    scan = read_values_subprocess(local_path, engine='openpyxl_scan')
                    scan_payload = scan.get('openpyxl_scan') if isinstance(scan, dict) else None
                    if scan_payload and getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                        print(f"   [openpyxl-scan:pre] sheets={len(scan_payload.get('sheets', []))}")
                    # 若掃描成功，並且值引擎可提供值，則直接組裝結果並返回
                    if scan_payload:
                        # 準備值引擎（子進程路徑）
                        values_by_sheet = {}
                        value_engine = getattr(settings, 'VALUE_ENGINE', 'polars')
                        try:
                            if value_engine in ('xml', 'polars_xml'):
                                if not silent:
                                    print(f"   [value-engine] {value_engine.upper()} via SUBPROCESS")
                                values_by_sheet = read_values_subprocess(local_path, engine=value_engine)
                            else:
                                values_by_sheet = {}
                        except Exception as _ve:
                            if not silent:
                                print(f"   [value-engine] subprocess pre-scan failed: {_ve}")
                            values_by_sheet = {}
                        # 外部參照映射（供 prettify 使用）
                        try:
                            ref_map = extract_external_refs(local_path)
                        except Exception:
                            ref_map = {}
                        result = {}
                        try:
                            sheets = (scan_payload or {}).get('sheets', [])
                            for s_idx, sh in enumerate(sheets, 1):
                                ws_title = sh.get('name') or f"Sheet{s_idx}"
                                formulas = sh.get('formula_cells', {}) or {}
                                ws_data = {}
                                for addr, f0 in formulas.items():
                                    external_ref = False
                                    f_before = str(f0) if f0 is not None else None
                                    f_pretty = f_before
                                    try:
                                        f_pretty = pretty_formula(f_before, ref_map=ref_map)
                                        import re as _re
                                        if f_before and _re.search(r"\[(\d+)\][^!\]]+!", f_before):
                                            external_ref = True
                                        elif isinstance(f_pretty, str) and _re.search(r"'[^']*\\\[[^\\\]]+\][^']*'!", f_pretty):
                                            external_ref = True
                                        elif isinstance(f_pretty, str) and _re.search(r"\[[^\]]+\][^!]+!", f_pretty):
                                            external_ref = True
                                    except Exception:
                                        pass
                                    vstr = None
                                    try:
                                        sheet_vals = (values_by_sheet or {}).get(ws_title) or {}
                                        vstr = sheet_vals.get(addr)
                                    except Exception:
                                        vstr = None
                                    cached_v = vstr if value_engine in ('polars', 'polars_xml', 'xml') else None
                                    ws_data[addr] = {
                                        'formula': f_pretty,
                                        'value': vstr,
                                        'cached_value': cached_v,
                                        'external_ref': bool(external_ref)
                                    }
                                if ws_data:
                                    result[ws_title] = ws_data
                        except Exception as _asm2:
                            if not silent:
                                print(f"   [openpyxl-scan:assemble-pre] failed: {_asm2}")
                        else:
                            if not silent:
                                print("   ✅ Excel 讀取完成（子進程 scan，主進程未載入 workbook）")
                            return result
        except Exception as _scan_pre:
            if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                print(f"   [openpyxl-scan:pre] skipped: {_scan_pre}")
        # 強制安全模式：FORMULA_ONLY_MODE 且嚴格時，掃描失敗即返回空結果，不回退主進程 openpyxl
        try:
            if bool(getattr(settings, 'FORMULA_ONLY_MODE', False)) and bool(getattr(settings, 'FORMULA_ONLY_STRICT_SAFE', False)):
                if not locals().get('scan_payload'):
                    if not silent:
                        print("   [safety] formula-only strict mode: scan failed -> return empty, skip main-process openpyxl")
                    return {}
        except Exception:
            pass
        # 多重保護機制：嘗試不同嘅載入方式
        wb = None
        load_attempts = [
            # 嘗試1：標準 read_only 模式
            {'read_only': True, 'data_only': False, 'keep_vba': False},
            # 嘗試2：關閉 VBA 支持
            {'read_only': True, 'data_only': False, 'keep_vba': False, 'keep_links': False},
            # 嘗試3：最保守模式
            {'read_only': True, 'data_only': True, 'keep_vba': False, 'keep_links': False},
        ]
        
        last_error = None
        for attempt_num, load_params in enumerate(load_attempts, 1):
            try:
                if not silent:
                    print(f"   🔄 載入嘗試 {attempt_num}/3: {load_params}")
                wb = safe_load_workbook(local_path, **load_params)
                if not silent:
                    print(f"   ✅ 載入成功 (嘗試 {attempt_num})")
                break
            except Exception as e:
                last_error = e
                if not silent:
                    print(f"   ❌ 載入失敗 (嘗試 {attempt_num}): {e}")
                if attempt_num < len(load_attempts):
                    continue
                else:
                    # 所有嘗試都失敗
                    if not silent:
                        print(f"   💀 所有載入方式都失敗，檔案可能損壞: {last_error}")
                    return None
        result = {}
        worksheet_count = len(wb.worksheets)
        if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
            print(f"   [xlsx-open-ok] sheets={worksheet_count}")

        if not silent and show_sheet_detail:
            print(f"   工作表數量: {worksheet_count}")

        # 解析外部參照映射，供 prettify 使用
        ref_map = extract_external_refs(local_path)

        # 可選：由子進程先掃描 openpyxl 公式/結構（避免主進程 iter_rows）
        try:
            from utils.xml_subproc import is_xml_subprocess_enabled, read_values_subprocess
            if is_xml_subprocess_enabled():
                scan = read_values_subprocess(local_path, engine='openpyxl_scan')
                scan_payload = scan.get('openpyxl_scan') if isinstance(scan, dict) else None
                if scan_payload and getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                    print(f"   [openpyxl-scan] sheets={len(scan_payload.get('sheets', []))}")
        except Exception as _scan_e:
            if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                print(f"   [openpyxl-scan] skipped: {_scan_e}")

        formula_cells_global = 0
        formula_coords_by_sheet = {}

        # 如果有來自子進程的 openpyxl 掃描結果，且為公式專注模式，則直接用掃描結果組裝，跳過主進程 iter_rows
        skip_iter_rows = False
        try:
            if (globals().get('scan_payload') or locals().get('scan_payload')) and bool(getattr(settings, 'FORMULA_ONLY_MODE', False)):
                payload = scan_payload  # 由上方子進程呼叫取得
                result = {}
                formula_coords_by_sheet = {}
                used = 0
                sheets = (payload or {}).get('sheets', [])
                for s_idx, sh in enumerate(sheets, 1):
                    ws_title = sh.get('name') or f"Sheet{s_idx}"
                    formulas = sh.get('formula_cells', {}) or {}
                    ws_data = {}
                    for addr, f0 in formulas.items():
                        external_ref = False
                        f_before = str(f0) if f0 is not None else None
                        f_pretty = f_before
                        try:
                            f_pretty = pretty_formula(f_before, ref_map=ref_map)
                            import re as _re
                            if f_before and _re.search(r"\[(\d+)\][^!\]]+!", f_before):
                                external_ref = True
                            elif isinstance(f_pretty, str) and _re.search(r"'[^']*\\\[[^\\\]]+\][^']*'!", f_pretty):
                                external_ref = True
                            elif isinstance(f_pretty, str) and _re.search(r"\[[^\]]+\][^!]+!", f_pretty):
                                external_ref = True
                        except Exception:
                            pass
                        # 從值引擎結果取值（名稱或索引對齊）
                        vstr = None
                        try:
                            sheet_vals = (values_by_sheet or {}).get(ws_title) or {}
                            if not sheet_vals and sheet_order and s_idx - 1 < len(sheet_order):
                                alt_name = sheet_order[s_idx - 1]
                                sheet_vals = (values_by_sheet or {}).get(alt_name) or {}
                            vstr = sheet_vals.get(addr)
                        except Exception:
                            vstr = None
                        cached_v = vstr if value_engine in ('polars', 'polars_xml', 'xml') else None
                        ws_data[addr] = {
                            'formula': f_pretty,
                            'value': vstr,
                            'cached_value': cached_v,
                            'external_ref': bool(external_ref)
                        }
                        used += 1
                    if ws_data:
                        result[ws_title] = ws_data
                        formula_coords_by_sheet[ws_title] = list(formulas.keys())
                formula_cells_global = sum(len(v) for v in formula_coords_by_sheet.values())
                skip_iter_rows = True
                if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                    print(f"   [openpyxl-scan] assembled cells={used}")
        except Exception as _asm_e:
            if not silent and getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                print(f"   [openpyxl-scan] assemble failed: {_asm_e}")

        # 準備值引擎
        value_engine = getattr(settings, 'VALUE_ENGINE', 'polars')
        persist_csv = bool(getattr(settings, 'CSV_PERSIST', False))
        persist_dir = getattr(settings, 'CACHE_FOLDER', None)
        values_by_sheet = {}

        try:
            if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                print(f"   [value-engine] selecting… pref={value_engine}")
            
            # 嘗試使用子進程讀取值（適用於 XML 相關引擎）
            subprocess_success = False
            if value_engine in ('xml', 'polars_xml'):
                try:
                    from utils.xml_subproc import is_xml_subprocess_enabled, read_values_subprocess
                    if is_xml_subprocess_enabled():
                        if not silent:
                            print(f"   [value-engine] {value_engine.upper()} via SUBPROCESS")
                        values_by_sheet = read_values_subprocess(local_path, engine=value_engine)
                        subprocess_success = True
                        if not silent:
                            print(f"   [value-engine] subprocess completed successfully")
                except Exception as e:
                    if not silent:
                        print(f"   [value-engine] subprocess failed: {e}")
                    # 半保守回退：嘗試非 XML 引擎（POLARS 或 PANDAS），嚴禁回退到主進程 XML
                    try:
                        from utils.value_engines.polars_reader import read_values_from_xlsx_via_polars
                        if not silent:
                            try:
                                import polars as _pl
                                import importlib
                                _x2c = importlib.util.find_spec('xlsx2csv') is not None
                                print(f"   [fallback] POLARS (polars={_pl.__version__}, xlsx2csv={'OK' if _x2c else 'NOT FOUND'}) | persist_csv={persist_csv}")
                            except Exception:
                                print("   [fallback] POLARS (version info unavailable)")
                        try:
                            values_by_sheet = read_values_from_xlsx_via_polars(local_path, persist_csv=persist_csv, persist_dir=persist_dir, sheet_count=len(wb.worksheets))
                        except TypeError:
                            values_by_sheet = read_values_from_xlsx_via_polars(local_path, persist_csv=persist_csv, persist_dir=persist_dir)
                        subprocess_success = True
                        value_engine = 'polars'
                    except Exception as e_polars:
                        if not silent:
                            print(f"   [fallback] polars path failed: {e_polars}")
                        try:
                            from utils.value_engines.pandas_reader import read_values_from_xlsx_via_pandas
                            if not silent:
                                print("   [fallback] PANDAS (via xlsx2csv -> pandas.read_csv)")
                            values_by_sheet = read_values_from_xlsx_via_pandas(local_path, persist_csv=persist_csv, persist_dir=persist_dir, sheet_count=len(wb.worksheets))
                            subprocess_success = True
                            value_engine = 'pandas'
                        except Exception as e_pandas:
                            if not silent:
                                print(f"   [fallback] pandas path failed: {e_pandas}")
                            # 保守方案：返回空結果並產生 Quest 報告
                            try:
                                from utils.enhanced_logging_and_error_handler import save_quest_report
                                sections = {
                                    "基本資訊": {
                                        "任務": "read_values",
                                        "檔案": local_path,
                                        "偏好引擎": getattr(settings, 'VALUE_ENGINE', 'xml'),
                                    },
                                    "錯誤": {
                                        "subprocess": str(e),
                                        "polars": str(e_polars),
                                        "pandas": str(e_pandas),
                                    }
                                }
                                qp = save_quest_report("值引擎回退失敗（已跳過檔案）", sections)
                                if not silent and qp:
                                    print(f"   [quest] 報告已生成: {qp}")
                            except Exception:
                                pass
                            values_by_sheet = {}
                            subprocess_success = True  # 標記為已處理（但為空），避免再走主進程 XML
            
            # 後續：如果上面沒有子進程或非 XML 引擎的結果，再根據用戶明確選擇處理
            if not subprocess_success:
                if value_engine == 'polars':
                    from utils.value_engines.polars_reader import read_values_from_xlsx_via_polars
                    if not silent:
                        try:
                            import polars as _pl
                            import importlib
                            _x2c = importlib.util.find_spec('xlsx2csv') is not None
                            print(f"   [value-engine] POLARS (polars={_pl.__version__}, xlsx2csv={'OK' if _x2c else 'NOT FOUND'}) | persist_csv={persist_csv}")
                        except Exception:
                            print("   [value-engine] POLARS (version info unavailable)")
                    try:
                        values_by_sheet = read_values_from_xlsx_via_polars(local_path, persist_csv=persist_csv, persist_dir=persist_dir, sheet_count=len(wb.worksheets))
                    except TypeError:
                        values_by_sheet = read_values_from_xlsx_via_polars(local_path, persist_csv=persist_csv, persist_dir=persist_dir)
                elif value_engine == 'pandas':
                    if not silent:
                        print("   [value-engine] PANDAS (via xlsx2csv -> pandas.read_csv)")
                    try:
                        from utils.value_engines.pandas_reader import read_values_from_xlsx_via_pandas
                        values_by_sheet = read_values_from_xlsx_via_pandas(local_path, persist_csv=persist_csv, persist_dir=persist_dir, sheet_count=len(wb.worksheets))
                    except Exception as e:
                        if not silent:
                            print(f"   [fallback->empty] pandas path failed: {e}")
                        values_by_sheet = {}
                else:
                    # 嚴禁回退到主進程 XML：直接返回空結果
                    if not silent:
                        print("   [safety] main-process XML fallback is disabled; returning empty result")
                    values_by_sheet = {}
        except Exception as e:
            # 不再回退到主進程 XML，僅記錄診斷與返回空
            try:
                import sys, importlib.util
                polars_ok = importlib.util.find_spec('polars') is not None
                x2c_ok = importlib.util.find_spec('xlsx2csv') is not None
                if not silent:
                    print(f"   [fallback-disabled] main-process XML fallback disabled | python={sys.executable} | polars={'OK' if polars_ok else 'NOT FOUND'} | xlsx2csv={'OK' if x2c_ok else 'NOT FOUND'} | err={e}")
            except Exception:
                pass
            values_by_sheet = {}

        # 值引擎返回的工作表 key（供對齊/診斷）
        try:
            sheet_order = list(values_by_sheet.keys())
            if not silent:
                print(f"   [value-engine] sheet keys from engine: {sheet_order}")
                if value_engine == 'polars_xml':
                    try:
                        for i, nm in enumerate(sheet_order, start=1):
                            vals = values_by_sheet.get(nm) or {}
                            cells = len(vals)
                            try:
                                nonempty = sum(1 for v in vals.values() if v not in (None, ''))
                            except Exception:
                                nonempty = cells
                            print(f"   [polars_xml] sheet {i} name='{nm}' cells={cells} nonempty={nonempty}")
                            if vals:
                                sample_items = list(vals.items())[:8]
                                sample_str = ', '.join([f"{k}->{repr(v)}" for k, v in sample_items])
                                print(f"   [polars_xml] sample: {sample_str}")
                    except Exception:
                        pass
        except Exception:
            sheet_order = []

        # 若值引擎未能返回任何工作表，遵循安全策略：不回退到主進程 XML
        if (not values_by_sheet) or (not sheet_order):
            if not silent:
                print("   [safety] no sheets from current engine; XML fallback disabled -> returning empty")
            values_by_sheet = {}
            sheet_order = []

        per_sheet_formula_provided = {}

        # 大文件保護設定
        LARGE_FILE_THRESHOLD = getattr(settings, 'LARGE_FILE_CELL_THRESHOLD', 1000000)  # 100萬個儲存格
        BATCH_SIZE = getattr(settings, 'EXCEL_BATCH_SIZE', 10000)  # 每批處理1萬行

        def process_cell_range(ws, ws_data, formula_addrs, min_row, max_row, min_col, max_col, 
                              sheet_vals, ref_map, value_engine, selected_key, per_sheet_formula_provided, silent):
            """處理指定範圍的儲存格"""
            def _col_to_letters(n: int) -> str:
                s = ''
                while n > 0:
                    n, r = divmod(n - 1, 26)
                    s = chr(65 + r) + s
                return s
            
            cell_count = 0
            formula_cells_count = 0
            
            for r_idx, row in enumerate(
                ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=False),
                start=min_row
            ):
                for c_idx, cell in enumerate(row, start=1):
                    addr = f"{_col_to_letters(c_idx)}{r_idx}"
                    external_ref = False
                    
                    try:
                        if hasattr(cell, 'formula') and cell.formula:
                            fstr = cell.formula
                        else:
                            fstr = get_cell_formula(cell)
                    except Exception:
                        fstr = None
                    
                    if fstr:
                        try:
                            s_before = str(fstr)
                            try:
                                fstr = pretty_formula(fstr, ref_map=ref_map)
                            except Exception:
                                pass
                            import re as _re
                            if _re.search(r"\[(\d+)\][^!\]]+!", s_before):
                                external_ref = True
                            elif isinstance(fstr, str) and _re.search(r"'[^']*\\\[[^\\\]]+\][^']*'!", fstr):
                                external_ref = True
                            elif isinstance(fstr, str) and _re.search(r"\[[^\]]+\][^!]+!", fstr):
                                external_ref = True
                            formula_addrs.append(addr)
                            formula_cells_count += 1
                        except Exception:
                            pass
                    
                    try:
                        vstr = sheet_vals.get(addr)
                    except Exception as _e:
                        if not silent:
                            print(f"   [read_error] sheet='{ws.title}' addr='{addr}' op='assemble' err={_e}")
                        try:
                            vstr = serialize_cell_value(getattr(cell, 'value', None))
                        except Exception:
                            vstr = None
                    
                    if fstr is not None or vstr is not None:
                        cached_v = vstr if value_engine in ('polars', 'polars_xml', 'xml') else None
                        ws_data[addr] = {
                            "formula": fstr,
                            "value": vstr,
                            "cached_value": cached_v,
                            "external_ref": bool(external_ref)
                        }
                        if fstr and (vstr is not None):
                            per_sheet_formula_provided[selected_key or ws.title] = per_sheet_formula_provided.get(selected_key or ws.title, 0) + 1
                        cell_count += 1
            
            return cell_count, formula_cells_count

        for idx, ws in enumerate(wb.worksheets, 1):
            if skip_iter_rows:
                break  # 已由 openpyxl 子進程掃描結果組裝，跳過主進程 iter_rows
            cell_count = 0
            ws_data = {}
            formula_addrs = []

            # 決定值引擎對應的 key
            selected_key = ws.title if ws.title in (values_by_sheet or {}) else None
            if selected_key is None and sheet_order:
                selected_key = list(values_by_sheet.keys())[idx - 1] if idx - 1 < len(values_by_sheet) else None
                if not silent and selected_key:
                    print(f"   [value-engine] sheet name mismatch: ws.title='{ws.title}' -> fallback to index key='{selected_key}'")

            sheet_vals = (values_by_sheet or {}).get(selected_key, {}) if selected_key else {}
            try:
                p_count = len(sheet_vals)
            except Exception:
                p_count = 0

            if not silent:
                try:
                    keys_list = list(sheet_vals.keys()) if isinstance(sheet_vals, dict) else []
                    show_keys = keys_list[:50]
                    if len(keys_list) > 50:
                        show_keys.append('...')
                except Exception:
                    show_keys = []
                try:
                    from utils.debug import debug_print
                    debug_print('map', f"ws_index={idx} ws_title='{ws.title}' -> key='{selected_key or ''}' provided={p_count}")
                    if show_keys:
                        debug_print('map', show_keys, label='keys', chunk=10)
                except Exception:
                    print(f"   [map] ws_index={idx} ws_title='{ws.title}' -> key='{selected_key or ''}' provided={p_count} keys={show_keys}")

            if ws.max_row >= 1 and ws.max_column >= 1:
                try:
                    # 大文件保護：檢查總儲存格數量
                    max_rows = ws.max_row
                    max_cols = ws.max_column
                    total_cells = max_rows * max_cols
                    
                    if not silent and total_cells > LARGE_FILE_THRESHOLD:
                        print(f"   ⚠️  大文件檢測: {total_cells:,} 個儲存格，啟用分批處理模式")
                    
                    if total_cells > LARGE_FILE_THRESHOLD:
                        # 分批處理大文件
                        for batch_start in range(1, max_rows + 1, BATCH_SIZE):
                            batch_end = min(batch_start + BATCH_SIZE - 1, max_rows)
                            if not silent:
                                print(f"   📦 處理批次: 行 {batch_start}-{batch_end} ({batch_end-batch_start+1:,} 行)")
                            
                            try:
                                from utils.gc_guard import gc_guard_any_thread
                                use_gc_guard = bool(getattr(settings, 'ENABLE_XML_GC_GUARD', True))
                                with gc_guard_any_thread(enabled=use_gc_guard, do_collect=False):
                                    batch_cells, batch_formulas = process_cell_range(
                                        ws, ws_data, formula_addrs, batch_start, batch_end, 1, max_cols,
                                        sheet_vals, ref_map, value_engine, selected_key, per_sheet_formula_provided, silent
                                    )
                                cell_count += batch_cells
                                formula_cells_global += batch_formulas
                                
                                #（移除強制 GC）交由 Python 自然回收，避免在 GC 階段觸發底層崩潰（Py3.11/3.12 + ET）
                                
                            except Exception as batch_e:
                                if not silent:
                                    print(f"   ❌ 批次處理失敗 (行 {batch_start}-{batch_end}): {batch_e}")
                                # 繼續處理下一批次
                                continue
                    else:
                        # 小文件：使用原有邏輯
                        from utils.gc_guard import gc_guard_any_thread
                        use_gc_guard = bool(getattr(settings, 'ENABLE_XML_GC_GUARD', True))
                        with gc_guard_any_thread(enabled=use_gc_guard, do_collect=False):
                            batch_cells, batch_formulas = process_cell_range(
                                ws, ws_data, formula_addrs, 1, max_rows, 1, max_cols,
                                sheet_vals, ref_map, value_engine, selected_key, per_sheet_formula_provided, silent
                            )
                        cell_count += batch_cells
                        formula_cells_global += batch_formulas
                        
                except Exception as _e:
                    if not silent:
                        print(f"   [read_error] sheet='{ws.title}' op='process_sheet' err={_e}")

            if show_sheet_detail and not silent:
                print(f"      處理工作表 {idx}/{worksheet_count}: {ws.title}（{cell_count} 有資料 cell）")

            if ws_data:
                result[ws.title] = ws_data
            if formula_addrs:
                formula_coords_by_sheet[ws.title] = formula_addrs

        # Phase 2：可選 cached value 比對（僅對公式格），避免外部參照刷新導致假變更
        try:
            if getattr(settings, 'ENABLE_FORMULA_VALUE_CHECK', False) and formula_cells_global > 0:
                # 若值引擎已提供 cached_value，則無需再做第二次 data_only pass
                provided = 0
                for sheet_name, coords in formula_coords_by_sheet.items():
                    for addr in coords:
                        if sheet_name in result and addr in result[sheet_name] and result[sheet_name][addr].get('cached_value') is not None:
                            provided += 1
                need_data_only = (provided == 0)
                if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                    try:
                        print(f"   [phase2] decision: enable={bool(getattr(settings,'ENABLE_FORMULA_VALUE_CHECK', False))} formulas={formula_cells_global} provided={provided} need_data_only={need_data_only}")
                    except Exception:
                        pass
                if not need_data_only:
                    if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                        try:
                            per_sheet_counts = {}
                            for sname, coords in formula_coords_by_sheet.items():
                                cnt = 0
                                for addr in coords:
                                    if sname in result and addr in result[sname] and result[sname][addr].get('cached_value') is not None:
                                        cnt += 1
                                per_sheet_counts[sname] = cnt
                            print(f"   [phase2] provided_from_value_engine total={provided} per_sheet={per_sheet_counts}")
                        except Exception:
                            print(f"   [phase2] 已由值引擎提供 cached value（{provided} 格），略過 openpyxl data_only 二次讀取。")
                else:
                    cap = int(getattr(settings, 'MAX_FORMULA_VALUE_CELLS', 50000))
                    if formula_cells_global > cap:
                        if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                            print(f"   ⏩ 公式格數量 {formula_cells_global} 超過上限 {cap}，略過值比對。")
                    else:
                        if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                            print(f"   讀取公式儲存格的 cached value（共 {formula_cells_global} 格）…")
                        # 使用子進程取得 data_only 值（避免主進程 openpyxl）
                        try:
                            from utils.xml_subproc import read_values_subprocess
                            # 先針對外部參照且缺值者補齊
                            external_missing = {}
                            if getattr(settings, 'ALWAYS_FETCH_VALUE_FOR_EXTERNAL_REFS', True):
                                cap2 = int(getattr(settings, 'EXTERNAL_REF_VALUE_FETCH_CAP', 0) or 0)
                                for sheet_name, coords in formula_coords_by_sheet.items():
                                    for addr in coords:
                                        try:
                                            meta = result.get(sheet_name, {}).get(addr, {})
                                            if meta and meta.get('external_ref') and (meta.get('cached_value') is None):
                                                if sheet_name not in external_missing:
                                                    external_missing[sheet_name] = []
                                                external_missing[sheet_name].append(addr)
                                        except Exception:
                                            pass
                            # 聚合所有需要補值的地址
                            coords_map = {}
                            per_sheet_added = {}
                            total_added = 0
                            for sheet_name, coords in formula_coords_by_sheet.items():
                                addrs = []
                                # 先外部參照缺值
                                addrs.extend(external_missing.get(sheet_name, []))
                                # 再其餘未有 cached_value 的地址
                                for addr in coords:
                                    if addr in addrs:
                                        continue
                                    if result.get(sheet_name, {}).get(addr, {}).get('cached_value') is None:
                                        addrs.append(addr)
                                if addrs:
                                    # cap 控制
                                    coords_map[sheet_name] = addrs
                            cap_total = int(getattr(settings, 'MAX_FORMULA_VALUE_CELLS', 50000))
                            # 統計 coords_map
                            try:
                                coords_total = sum(len(v) for v in coords_map.values())
                                print(f"   [phase2-subproc] coords_map total={coords_total} sheets={len(coords_map)} cap={cap_total}")
                            except Exception:
                                pass
                            # 呼叫子進程 data_only 取值（使用專用 API）
                            try:
                                from utils.xml_subproc import read_data_only_values_subprocess
                                dov = read_data_only_values_subprocess(local_path, coords_map, cap_total)
                            except Exception as _api_e:
                                # 後備：用通用 read_values_subprocess，但它不帶座標，回傳可能為空
                                try:
                                    data_only_res = read_values_subprocess(local_path, engine='data_only_values')
                                    dov = data_only_res.get('data_only_values', {}) if isinstance(data_only_res, dict) else {}
                                except Exception:
                                    dov = {}
                            # 統計回傳總量
                            try:
                                dov_total = sum(len(v) for v in (dov or {}).values())
                                print(f"   [phase2-subproc] data_only return total={dov_total}")
                            except Exception:
                                pass
                            # 合併結果
                            for sheet_name, addrs in coords_map.items():
                                for addr in addrs:
                                    try:
                                        sval = (dov.get(sheet_name, {}) if isinstance(dov, dict) else {}).get(addr)
                                    except Exception:
                                        sval = None
                                    if sheet_name in result and addr in result[sheet_name]:
                                        before = result[sheet_name][addr].get('cached_value')
                                        result[sheet_name][addr]['cached_value'] = sval
                                        if sval is not None and before is None:
                                            per_sheet_added[sheet_name] = per_sheet_added.get(sheet_name, 0) + 1
                                            total_added += 1
                            if getattr(settings, 'SHOW_DEBUG_MESSAGES', False):
                                try:
                                    print(f"   [phase2-subproc] data_only fetched cells total={total_added}")
                                    if per_sheet_added:
                                        items = sorted(per_sheet_added.items(), key=lambda x: x[0])
                                        print("   [phase2-subproc] per_sheet added:")
                                        for nm, cnt in items:
                                            print(f"      - {nm}: {cnt}")
                                except Exception:
                                    pass
                        except Exception as sub_e:
                            logging.warning(f"子進程 data_only 取值失敗：{sub_e}")
        except Exception as e:
            logging.warning(f"讀取 cached value 失敗：{e}")

        try:
            wb.close()
            wb = None
        except Exception:
            pass

        if not silent and show_sheet_detail:
            print(f"   ✅ Excel 讀取完成")

        return result

    except Exception as e:
        if not silent:
            logging.error(f"Excel 讀取失敗: {e}")
        return None
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
            del wb

        # 重置全局變數
        settings.current_processing_file = None
        settings.processing_start_time = None

def hash_excel_content(cells_dict):
    """
    計算 Excel 內容的雜湊值
    """
    if cells_dict is None: 
        return None
    
    try:
        content_str = json.dumps(cells_dict, sort_keys=True, ensure_ascii=False)
        return hashlib.md5(content_str.encode('utf-8')).hexdigest()
    except (TypeError, json.JSONEncodeError) as e:
        logging.error(f"計算 Excel 內容雜湊值失敗: {e}")

        return None
