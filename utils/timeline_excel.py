import os
from datetime import datetime
import config.settings as settings

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except Exception:  # 若環境缺 openpyxl，先容錯（主流程不可因 timeline 失敗而中斷）
    Workbook = None
    load_workbook = None

TIMELINE_DIR = None
TIMELINE_XLSX = None


def _init_paths():
    global TIMELINE_DIR, TIMELINE_XLSX
    if TIMELINE_DIR is None:
        TIMELINE_DIR = os.path.join(getattr(settings, 'LOG_FOLDER', '.'), 'timeline')
    if TIMELINE_XLSX is None:
        TIMELINE_XLSX = os.path.join(TIMELINE_DIR, 'Timeline.xlsx')
    os.makedirs(TIMELINE_DIR, exist_ok=True)


def _autosize(ws):
    try:
        for col in ws.columns:
            max_len = 8
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    v = str(cell.value) if cell.value is not None else ''
                    if len(v) > max_len:
                        max_len = len(v)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    except Exception:
        pass


def export_event(event: dict):
    """
    Append a timeline event to Timeline.xlsx.
    Columns: 時間, 檔名, 工作表, 變更數, 作者, 事件#, 快照, 詳表, 路徑
    """
    try:
        _init_paths()
        ts = event.get('timestamp') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_path = event.get('file') or ''
        filename = event.get('filename') or os.path.basename(file_path)
        ws_name = event.get('worksheet') or 'all'
        changes = event.get('changes') or 0
        author = event.get('author') or ''
        evtno = event.get('event_number') or ''
        snap = event.get('snapshot_path') or ''
        per_event = event.get('per_event_path') or ''
        row = [ts, filename, ws_name, changes, author, evtno, snap, per_event, file_path]
        if Workbook is None:
            # Fallback: write CSV when openpyxl not available
            csv_path = os.path.join(TIMELINE_DIR, 'Timeline.csv')
            import csv
            header = ['時間','檔名','工作表','變更數','作者','事件#','快照','詳表','路徑']
            write_header = not os.path.exists(csv_path)
            with open(csv_path, 'a', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f)
                if write_header:
                    w.writerow(header)
                w.writerow(row)
            return
        # Excel path
        if os.path.exists(TIMELINE_XLSX):
            try:
                wb = load_workbook(TIMELINE_XLSX)
            except Exception:
                wb = Workbook()
        else:
            wb = Workbook()
        ws = wb.active
        ws.title = 'Events'
        # Header
        if ws.max_row == 1 and (ws.max_column == 1 and ws.cell(1,1).value is None):
            headers = ['時間','檔名','工作表','變更數','作者','事件#','快照','詳表','路徑']
            ws.append(headers)
        ws.append(row)
        # Hyperlinks
        try:
            if snap:
                ws.cell(ws.max_row, 7).hyperlink = snap
                ws.cell(ws.max_row, 7).style = 'Hyperlink'
        except Exception:
            pass
        try:
            if per_event:
                ws.cell(ws.max_row, 8).hyperlink = per_event
                ws.cell(ws.max_row, 8).style = 'Hyperlink'
        except Exception:
            pass
        _autosize(ws)
        wb.save(TIMELINE_XLSX)
    except Exception:
        # timeline 失敗不可影響主流程
        pass
