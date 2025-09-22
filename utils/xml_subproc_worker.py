"""
XML 子進程工作腳本
在獨立子進程中執行 XML 解析任務，避免崩潰影響主程式
"""
import os
import sys
import json
import traceback
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, Any, Optional


def debug_print(message: str, worker_id: int = 0):
    """輸出 debug 訊息到 stderr（避免與結果輸出混淆）"""
    print(f"[xml-worker-{worker_id}] {message}", file=sys.stderr, flush=True)


def extract_external_refs_task(file_path: str, safe_mode: bool = False, worker_id: int = 0) -> Dict[int, str]:
    """
    提取 Excel 外部參照
    
    Args:
        file_path: Excel 檔案路徑
        safe_mode: 是否使用安全模式
        worker_id: 工作者 ID
        
    Returns:
        外部參照映射 {index: path}
    """
    debug_print(f"extract_external_refs start file={os.path.basename(file_path)} safe_mode={safe_mode}", worker_id)
    
    ref_map = {}
    
    try:
        # 安全模式：使用更保守的設定
        if safe_mode:
            debug_print("using safe_mode: conservative XML parsing", worker_id)
            # 可以在這裡添加 lxml 的 recover 模式等
        
        with zipfile.ZipFile(file_path, 'r') as z:
            # 讀取 workbook.xml.rels 找外部連結
            try:
                rels_xml = z.read('xl/_rels/workbook.xml.rels')
                rels = ET.fromstring(rels_xml)
                
                external_links = []
                for rel in rels.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    if rel.attrib.get('Type', '').endswith('/externalLink'):
                        target = rel.attrib.get('Target', '')  # e.g., externalLinks/externalLink1.xml
                        external_links.append(target)
                
                debug_print(f"found {len(external_links)} external link files", worker_id)
                
                # 解析每個外部連結檔案
                for target in external_links:
                    import re
                    m = re.search(r'externalLink(\d+)\.xml', target)
                    if not m:
                        continue
                    
                    num = int(m.group(1))
                    path = ''
                    
                    # 方法 1：從 externalLinkN.xml 的 externalBookPr@href 取得路徑
                    try:
                        link_xml = z.read(f'xl/{target}')
                        link_tree = ET.fromstring(link_xml)
                        book_elem = link_tree.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}externalBookPr')
                        if book_elem is not None:
                            path = book_elem.attrib.get('href', '')
                            debug_print(f"external_link_{num} method1 path={path}", worker_id)
                    except Exception as e:
                        debug_print(f"external_link_{num} method1 failed: {e}", worker_id)
                    
                    # 方法 2：從 _rels/externalLinkN.xml.rels 取得路徑
                    if not path:
                        try:
                            rels_path = f"xl/externalLinks/_rels/externalLink{num}.xml.rels"
                            if rels_path in z.namelist():
                                link_rels_xml = z.read(rels_path)
                                link_rels = ET.fromstring(link_rels_xml)
                                rel_node = link_rels.find('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
                                if rel_node is not None:
                                    path = rel_node.attrib.get('Target', '')
                                    debug_print(f"external_link_{num} method2 path={path}", worker_id)
                        except Exception as e:
                            debug_print(f"external_link_{num} method2 failed: {e}", worker_id)
                    
                    ref_map[num] = path or ''
                
            except Exception as e:
                debug_print(f"extract_external_refs failed: {e}", worker_id)
                if safe_mode:
                    # 安全模式下，即使失敗也返回空結果而不是拋出異常
                    debug_print("safe_mode: returning empty result instead of raising exception", worker_id)
                    return {}
                else:
                    raise
    
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as e:
        debug_print(f"extract_external_refs error: {type(e).__name__}: {e}", worker_id)
        if safe_mode:
            debug_print("safe_mode: returning empty result for parsing error", worker_id)
            return {}
        else:
            raise RuntimeError(f"XML 解析失敗: {e}")
    
    debug_print(f"extract_external_refs completed refs_count={len(ref_map)}", worker_id)
    return ref_map


def read_values_task(file_path: str, engine: str = 'xml', safe_mode: bool = False, worker_id: int = 0) -> Dict[str, Dict[str, Any]]:
    """
    讀取 Excel 儲存格值
    
    Args:
        file_path: Excel 檔案路徑
        engine: 值引擎類型
        safe_mode: 是否使用安全模式
        worker_id: 工作者 ID
        
    Returns:
        工作表資料 {sheet_name: {address: value}}
    """
    debug_print(f"read_values start file={os.path.basename(file_path)} engine={engine} safe_mode={safe_mode}", worker_id)
    
    try:
        if engine == 'xml':
            return _read_values_xml(file_path, safe_mode, worker_id)
        elif engine == 'polars_xml':
            return _read_values_polars_xml(file_path, safe_mode, worker_id)
        else:
            raise ValueError(f"不支援的值引擎: {engine}")
            
    except Exception as e:
        debug_print(f"read_values error: {type(e).__name__}: {e}", worker_id)
        if safe_mode:
            debug_print("safe_mode: returning empty result for read error", worker_id)
            return {}
        else:
            raise RuntimeError(f"讀取值失敗: {e}")


def _scan_openpyxl_formulas(file_path: str, worker_id: int, safe_mode: bool = False) -> Dict[str, Any]:
    """使用 openpyxl 載入活頁簿並掃描所有工作表的公式（僅回傳公式，避免大量值資料）。"""
    from openpyxl import load_workbook as _lw
    debug_print(f"scan_openpyxl start file={os.path.basename(file_path)} safe_mode={safe_mode}", worker_id)
    try:
        wb = _lw(file_path, read_only=True, data_only=False)
    except Exception as e:
        if safe_mode:
            debug_print(f"scan_openpyxl load_workbook failed (safe): {e}", worker_id)
            return {"sheets": []}
        raise
    out = {"sheets": []}
    try:
        for ws in wb.worksheets:
            try:
                max_row = getattr(ws, 'max_row', 0) or 0
                max_col = getattr(ws, 'max_column', 0) or 0
            except Exception:
                max_row = 0
                max_col = 0
            formulas: Dict[str, str] = {}
            # 逐行掃描，僅提取公式（避免傳輸大量值）
            try:
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        try:
                            if cell is None:
                                continue
                            # 取地址
                            addr = getattr(cell, 'coordinate', None)
                            if not addr:
                                continue
                            # 判斷是否公式
                            fstr = None
                            if hasattr(cell, 'formula') and cell.formula:
                                fstr = cell.formula
                            elif getattr(cell, 'data_type', None) == 'f':
                                val = getattr(cell, 'value', None)
                                try:
                                    from openpyxl.worksheet.formula import ArrayFormula as _AF
                                    if isinstance(val, _AF):
                                        fstr = val.text if hasattr(val, 'text') else str(val)
                                    else:
                                        fstr = val
                                except Exception:
                                    fstr = val
                            if fstr:
                                formulas[addr] = str(fstr)
                        except Exception:
                            continue
            except Exception as e:
                debug_print(f"scan_openpyxl iter_rows failed on sheet '{ws.title}': {e}", worker_id)
            out["sheets"].append({
                "name": ws.title,
                "max_row": int(max_row),
                "max_col": int(max_col),
                "formula_cells": formulas,
            })
    finally:
        try:
            wb.close()
        except Exception:
            pass
    debug_print(f"scan_openpyxl completed sheets={len(out['sheets'])}", worker_id)
    return out


def _serialize_value(v):
    try:
        from datetime import datetime
        if v is None:
            return None
        if isinstance(v, (int, float, bool, str)):
            return v
        if isinstance(v, datetime):
            return v.isoformat()
        return str(v)
    except Exception:
        return None


def _fetch_data_only_values(file_path: str, coords_by_sheet: Dict[str, Any], worker_id: int, safe_mode: bool = False, cap: int = 0) -> Dict[str, Any]:
    """以 data_only 模式讀取指定座標的值，僅返回所需位址的值。"""
    from openpyxl import load_workbook as _lw
    debug_print(f"data_only_values start file={os.path.basename(file_path)} cap={cap} safe_mode={safe_mode}", worker_id)
    out: Dict[str, Dict[str, Any]] = {}
    try:
        wb = _lw(file_path, read_only=True, data_only=True)
    except Exception as e:
        if safe_mode:
            debug_print(f"data_only_values load_workbook failed (safe): {e}", worker_id)
            return out
        raise
    try:
        total = 0
        for sheet_name, addrs in (coords_by_sheet or {}).items():
            try:
                ws = wb[sheet_name]
            except Exception:
                continue
            out_sheet = {}
            for addr in (addrs or []):
                if cap and total >= cap:
                    break
                try:
                    val = ws[addr].value
                except Exception:
                    val = None
                out_sheet[addr] = _serialize_value(val)
                total += 1
            if out_sheet:
                out[sheet_name] = out_sheet
    finally:
        try:
            wb.close()
        except Exception:
            pass
    debug_print(f"data_only_values completed cells={sum(len(v) for v in out.values())}", worker_id)
    return out


def _read_values_xml(file_path: str, safe_mode: bool, worker_id: int) -> Dict[str, Dict[str, Any]]:
    """使用純 XML 方式讀取值"""
    debug_print("using xml engine", worker_id)
    
    with zipfile.ZipFile(file_path, 'r') as z:
        # 建立共享字串表
        shared_strings = []
        try:
            if 'xl/sharedStrings.xml' in z.namelist():
                ss_xml = z.read('xl/sharedStrings.xml')
                
                if safe_mode:
                    # 安全模式：使用更保守的 XML 解析
                    debug_print("safe_mode: using conservative XML parsing for sharedStrings", worker_id)
                
                root = ET.fromstring(ss_xml)
                ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                
                for si in root.findall('a:si', ns):
                    text_parts = []
                    for t in si.findall('.//a:t', ns):
                        text_parts.append(t.text or '')
                    shared_strings.append(''.join(text_parts))
                
                debug_print(f"loaded {len(shared_strings)} shared strings", worker_id)
        except Exception as e:
            debug_print(f"shared_strings load failed: {e}", worker_id)
            shared_strings = []
        
        # 取得工作表名稱
        sheet_names = []
        try:
            wb_xml = z.read('xl/workbook.xml')
            wroot = ET.fromstring(wb_xml)
            ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            for s in wroot.findall('a:sheets/a:sheet', ns):
                sheet_names.append(s.attrib.get('name', f'Sheet{len(sheet_names)+1}'))
            
            debug_print(f"found {len(sheet_names)} sheets: {sheet_names}", worker_id)
        except Exception as e:
            debug_print(f"workbook.xml parse failed: {e}", worker_id)
            sheet_names = []
        
        # 讀取每個工作表的資料
        result = {}
        for idx, name in enumerate(sheet_names, start=1):
            sheet_path = f'xl/worksheets/sheet{idx}.xml'
            if sheet_path not in z.namelist():
                continue
            
            try:
                xml = z.read(sheet_path)
                root = ET.fromstring(xml)
                ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                
                values = {}
                cell_count = 0
                
                for c in root.findall('.//a:c', ns):
                    addr = c.attrib.get('r')
                    if not addr:
                        continue
                    
                    t = c.attrib.get('t')  # 類型
                    v_node = c.find('a:v', ns)
                    if v_node is None:
                        continue
                    
                    raw = v_node.text
                    if raw is None:
                        val = None
                    elif t == 's':  # 共享字串
                        try:
                            idx = int(raw)
                            val = shared_strings[idx] if 0 <= idx < len(shared_strings) else ''
                        except (ValueError, IndexError):
                            val = ''
                    elif t == 'b':  # 布林值
                        val = 'TRUE' if raw in ('1', 'true', 'TRUE') else 'FALSE'
                    else:  # 數字或一般
                        val = raw
                    
                    values[addr] = val
                    cell_count += 1
                
                result[name] = values
                debug_print(f"sheet '{name}' loaded {cell_count} cells", worker_id)
                
            except Exception as e:
                debug_print(f"sheet '{name}' parse failed: {e}", worker_id)
                if not safe_mode:
                    raise
                # 安全模式下跳過失敗的工作表
                continue
        
        debug_print(f"read_values_xml completed sheets={len(result)}", worker_id)
        return result


def _read_values_polars_xml(file_path: str, safe_mode: bool, worker_id: int) -> Dict[str, Dict[str, Any]]:
    """使用 polars_xml 引擎讀取值（簡化版）"""
    debug_print("using polars_xml engine", worker_id)
    
    # 這裡可以調用實際的 polars_xml_reader 邏輯
    # 暫時使用 XML 引擎作為後備
    debug_print("polars_xml not implemented, falling back to xml", worker_id)
    return _read_values_xml(file_path, safe_mode, worker_id)


def execute_task(task_input: Dict[str, Any]) -> Dict[str, Any]:

    """
    執行任務的主要入口點
    
    Args:
        task_input: 任務輸入資料
        
    Returns:
        任務結果
    """
    task_type = task_input.get('task_type')
    task_data = task_input.get('task_data', {})
    safe_mode = task_input.get('safe_mode', False)
    worker_id = task_input.get('worker_id', 0)
    
    debug_print(f"execute_task start type={task_type} safe_mode={safe_mode}", worker_id)
    
    try:
        if task_type == 'extract_refs':
            file_path = task_data['file_path']
            external_refs = extract_external_refs_task(file_path, safe_mode, worker_id)
            return {
                'success': True,
                'external_refs': external_refs,
                'worker_id': worker_id
            }
        
        elif task_type == 'read_values':
            file_path = task_data['file_path']
            engine = task_data.get('engine', 'xml')
            # 若 engine 指定為 'openpyxl_scan'，只回傳公式掃描（為主進程提供結構/公式資訊）
            if engine == 'openpyxl_scan':
                scan = _scan_openpyxl_formulas(file_path, worker_id, safe_mode)
                return {
                    'success': True,
                    'openpyxl_scan': scan,
                    'worker_id': worker_id
                }
            # 若 engine 指定為 'data_only_values'，則對指定座標取值
            if engine == 'data_only_values':
                coords = task_data.get('coords_by_sheet', {}) or {}
                cap = int(task_data.get('cap', 0) or 0)
                data_vals = _fetch_data_only_values(file_path, coords, worker_id, safe_mode, cap)
                return {
                    'success': True,
                    'data_only_values': data_vals,
                    'worker_id': worker_id
                }
            values_by_sheet = read_values_task(file_path, engine, safe_mode, worker_id)
            return {
                'success': True,
                'values_by_sheet': values_by_sheet,
                'worker_id': worker_id
            }
        
        elif task_type == 'read_meta':
            # 讀取 metadata（例如最後作者）
            file_path = task_data['file_path']
            debug_print("read_meta start", worker_id)
            last_author = None
            try:
                import zipfile, xml.etree.ElementTree as ET
                with zipfile.ZipFile(file_path, 'r') as z:
                    core_xml = z.read('docProps/core.xml')
                    root = ET.fromstring(core_xml)
                    ns = {
                        'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
                        'dc': 'http://purl.org/dc/elements/1.1/'
                    }
                    node = root.find('cp:lastModifiedBy', ns)
                    if node is None:
                        node = root.find('dc:lastModifiedBy', ns)
                    last_author = (node.text or '').strip() if node is not None else None
            except Exception as me:
                debug_print(f"read_meta failed: {me}", worker_id)
                if not safe_mode:
                    raise
                # 安全模式下返回空 meta
                last_author = None
            return {
                'success': True,
                'meta': {
                    'last_author': last_author
                },
                'worker_id': worker_id
            }
        
        else:
            raise ValueError(f"不支援的任務類型: {task_type}")
    
    except Exception as e:
        debug_print(f"execute_task failed: {type(e).__name__}: {e}", worker_id)
        return {
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__,
            'worker_id': worker_id
        }


def main():
    """子進程主要入口點"""
    try:
        # 強制 I/O 編碼為 UTF-8，避免父進程解碼錯誤
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='strict')  # 僅允許 UTF-8，出錯即報
        except Exception:
            pass
        try:
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')  # 調試輸出容忍非 UTF-8
        except Exception:
            pass
        
        # 從 stdin 讀取任務輸入
        input_data = sys.stdin.read()
        task_input = json.loads(input_data)
        
        # 執行任務
        result = execute_task(task_input)
        
        # 輸出結果到 stdout（ASCII 僅，避免編碼環境干擾）
        output = json.dumps(result, ensure_ascii=True, indent=None, separators=(',', ':'))
        print(output, flush=True)
        
        # 成功退出
        sys.exit(0)
        
    except json.JSONDecodeError as e:
        error_result = {
            'success': False,
            'error': f'JSON 解析失敗: {e}',
            'error_type': 'JSONDecodeError'
        }
        print(json.dumps(error_result, ensure_ascii=False), flush=True)
        sys.exit(1)
        
    except Exception as e:
        error_result = {
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__,
            'traceback': traceback.format_exc()
        }
        print(json.dumps(error_result, ensure_ascii=False), flush=True)
        sys.exit(1)


if __name__ == '__main__':
    main()