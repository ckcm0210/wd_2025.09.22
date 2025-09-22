"""
Excel 相關的子進程任務
將所有 openpyxl 操作隔離到子進程中執行
"""
import os
import sys
import json
from typing import Dict, Any, List, Optional

def debug_print(message: str, worker_id: int = 0):
    """輸出 debug 訊息到 stderr"""
    print(f"[excel-worker-{worker_id}] {message}", file=sys.stderr, flush=True)

def safe_load_workbook_task(file_path: str, safe_mode: bool = False, worker_id: int = 0, **kwargs):
    """
    安全載入 Excel 檔案 (子進程版本)
    
    Args:
        file_path: Excel 檔案路徑
        safe_mode: 是否使用安全模式
        worker_id: 工作者 ID
        **kwargs: openpyxl load_workbook 參數
        
    Returns:
        工作簿基本資訊和工作表列表
    """
    debug_print(f"safe_load_workbook start file={os.path.basename(file_path)} safe_mode={safe_mode}", worker_id)
    
    try:
        from openpyxl import load_workbook
        
        # 預設安全參數
        default_params = {
            'read_only': True,
            'data_only': False,
            'keep_vba': False,
            'keep_links': False
        }
        
        # 合併用戶參數
        load_params = {**default_params, **kwargs}
        
        if safe_mode:
            # 安全模式：更保守的設定
            load_params.update({
                'read_only': True,
                'data_only': True,
                'keep_vba': False,
                'keep_links': False
            })
            debug_print("using safe_mode: conservative openpyxl settings", worker_id)
        
        debug_print(f"load_params: {load_params}", worker_id)
        
        # 載入工作簿
        wb = load_workbook(file_path, **load_params)
        
        # 提取基本資訊
        workbook_info = {
            'filename': os.path.basename(file_path),
            'worksheet_count': len(wb.worksheets),
            'worksheets': [],
            'load_params': load_params
        }
        
        # 提取工作表資訊
        for idx, ws in enumerate(wb.worksheets):
            try:
                max_row = getattr(ws, 'max_row', 0) or 0
                max_col = getattr(ws, 'max_column', 0) or 0
                
                sheet_info = {
                    'index': idx,
                    'name': ws.title,
                    'max_row': max_row,
                    'max_col': max_col,
                    'total_cells': max_row * max_col if max_row and max_col else 0
                }
                workbook_info['worksheets'].append(sheet_info)
                
            except Exception as e:
                debug_print(f"worksheet '{ws.title}' info extraction failed: {e}", worker_id)
                if not safe_mode:
                    raise
                # 安全模式下跳過有問題的工作表
                continue
        
        wb.close()
        debug_print(f"safe_load_workbook completed sheets={len(workbook_info['worksheets'])}", worker_id)
        return workbook_info
        
    except Exception as e:
        debug_print(f"safe_load_workbook failed: {type(e).__name__}: {e}", worker_id)
        if safe_mode:
            debug_print("safe_mode: returning minimal result", worker_id)
            return {
                'filename': os.path.basename(file_path),
                'worksheet_count': 0,
                'worksheets': [],
                'error': str(e)
            }
        else:
            raise RuntimeError(f"Excel 載入失敗: {e}")

def extract_all_formulas_task(file_path: str, safe_mode: bool = False, worker_id: int = 0, 
                             batch_size: int = 10000) -> Dict[str, Any]:
    """
    提取所有工作表的公式 (子進程版本)
    
    Args:
        file_path: Excel 檔案路徑
        safe_mode: 是否使用安全模式
        worker_id: 工作者 ID
        batch_size: 批次大小 (防止記憶體溢出)
        
    Returns:
        所有工作表的公式資料
    """
    debug_print(f"extract_all_formulas start file={os.path.basename(file_path)} batch_size={batch_size}", worker_id)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.formula import ArrayFormula
        
        # 載入工作簿 (只讀模式，不載入數據)
        wb = load_workbook(file_path, read_only=True, data_only=False, keep_vba=False)
        
        result = {
            'filename': os.path.basename(file_path),
            'worksheets': []
        }
        
        for ws_idx, ws in enumerate(wb.worksheets):
            debug_print(f"processing worksheet {ws_idx+1}/{len(wb.worksheets)}: '{ws.title}'", worker_id)
            
            try:
                max_row = getattr(ws, 'max_row', 0) or 0
                max_col = getattr(ws, 'max_column', 0) or 0
                
                worksheet_data = {
                    'name': ws.title,
                    'max_row': max_row,
                    'max_col': max_col,
                    'formulas': {},
                    'formula_count': 0
                }
                
                # 分批處理大文件
                if max_row * max_col > batch_size:
                    debug_print(f"large worksheet detected ({max_row}x{max_col}), using batch processing", worker_id)
                    
                    for batch_start in range(1, max_row + 1, batch_size // max_col if max_col > 0 else batch_size):
                        batch_end = min(batch_start + (batch_size // max_col if max_col > 0 else batch_size) - 1, max_row)
                        
                        try:
                            batch_formulas = _extract_formulas_from_range(
                                ws, batch_start, batch_end, 1, max_col, worker_id, safe_mode
                            )
                            worksheet_data['formulas'].update(batch_formulas)
                            
                        except Exception as batch_e:
                            debug_print(f"batch {batch_start}-{batch_end} failed: {batch_e}", worker_id)
                            if not safe_mode:
                                raise
                            continue
                else:
                    # 小文件：直接處理
                    worksheet_data['formulas'] = _extract_formulas_from_range(
                        ws, 1, max_row, 1, max_col, worker_id, safe_mode
                    )
                
                worksheet_data['formula_count'] = len(worksheet_data['formulas'])
                result['worksheets'].append(worksheet_data)
                
                debug_print(f"worksheet '{ws.title}' completed, {worksheet_data['formula_count']} formulas", worker_id)
                
            except Exception as ws_e:
                debug_print(f"worksheet '{ws.title}' failed: {ws_e}", worker_id)
                if not safe_mode:
                    raise
                continue
        
        wb.close()
        
        total_formulas = sum(ws['formula_count'] for ws in result['worksheets'])
        debug_print(f"extract_all_formulas completed, total formulas: {total_formulas}", worker_id)
        
        return result
        
    except Exception as e:
        debug_print(f"extract_all_formulas failed: {type(e).__name__}: {e}", worker_id)
        if safe_mode:
            return {
                'filename': os.path.basename(file_path),
                'worksheets': [],
                'error': str(e)
            }
        else:
            raise RuntimeError(f"公式提取失敗: {e}")

def _extract_formulas_from_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, 
                                worker_id: int, safe_mode: bool) -> Dict[str, str]:
    """從指定範圍提取公式"""
    formulas = {}
    
    try:
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, 
                               min_col=min_col, max_col=max_col, values_only=False):
            for cell in row:
                if cell is None:
                    continue
                
                try:
                    # 獲取儲存格地址
                    addr = getattr(cell, 'coordinate', None)
                    if not addr:
                        continue
                    
                    # 檢查是否為公式
                    formula_text = None
                    
                    if hasattr(cell, 'formula') and cell.formula:
                        formula_text = cell.formula
                    elif getattr(cell, 'data_type', None) == 'f':
                        val = getattr(cell, 'value', None)
                        if val is not None:
                            try:
                                from openpyxl.worksheet.formula import ArrayFormula
                                if isinstance(val, ArrayFormula):
                                    formula_text = val.text if hasattr(val, 'text') else str(val)
                                else:
                                    formula_text = str(val)
                            except Exception:
                                formula_text = str(val)
                    
                    if formula_text:
                        formulas[addr] = str(formula_text)
                        
                except Exception as cell_e:
                    if not safe_mode:
                        debug_print(f"cell processing failed: {cell_e}", worker_id)
                    continue
                    
    except Exception as range_e:
        debug_print(f"range processing failed: {range_e}", worker_id)
        if not safe_mode:
            raise
    
    return formulas

def extract_cell_values_task(file_path: str, target_addresses: Dict[str, List[str]], 
                           safe_mode: bool = False, worker_id: int = 0, 
                           use_data_only: bool = True) -> Dict[str, Dict[str, Any]]:
    """
    提取指定儲存格的值 (子進程版本)
    
    Args:
        file_path: Excel 檔案路徑
        target_addresses: {worksheet_name: [address_list]} 要提取的儲存格地址
        safe_mode: 是否使用安全模式
        worker_id: 工作者 ID
        use_data_only: 是否使用 data_only 模式
        
    Returns:
        儲存格值資料 {worksheet_name: {address: value}}
    """
    debug_print(f"extract_cell_values start file={os.path.basename(file_path)} data_only={use_data_only}", worker_id)
    
    try:
        from openpyxl import load_workbook
        from datetime import datetime
        
        # 載入工作簿
        wb = load_workbook(file_path, read_only=True, data_only=use_data_only, keep_vba=False)
        
        result = {}
        total_cells = sum(len(addrs) for addrs in target_addresses.values())
        processed_cells = 0
        
        debug_print(f"target cells: {total_cells}", worker_id)
        
        for ws_name, addresses in target_addresses.items():
            if not addresses:
                continue
                
            try:
                ws = wb[ws_name]
                ws_values = {}
                
                for addr in addresses:
                    try:
                        cell = ws[addr]
                        value = cell.value
                        
                        # 序列化值
                        if value is None:
                            serialized_value = None
                        elif isinstance(value, (int, float, bool, str)):
                            serialized_value = value
                        elif isinstance(value, datetime):
                            serialized_value = value.isoformat()
                        else:
                            serialized_value = str(value)
                        
                        ws_values[addr] = serialized_value
                        processed_cells += 1
                        
                    except Exception as cell_e:
                        debug_print(f"cell {addr} failed: {cell_e}", worker_id)
                        if not safe_mode:
                            raise
                        ws_values[addr] = None
                        continue
                
                if ws_values:
                    result[ws_name] = ws_values
                    
                debug_print(f"worksheet '{ws_name}' completed, {len(ws_values)} cells", worker_id)
                
            except KeyError:
                debug_print(f"worksheet '{ws_name}' not found", worker_id)
                if not safe_mode:
                    raise
                continue
            except Exception as ws_e:
                debug_print(f"worksheet '{ws_name}' failed: {ws_e}", worker_id)
                if not safe_mode:
                    raise
                continue
        
        wb.close()
        debug_print(f"extract_cell_values completed, {processed_cells}/{total_cells} cells", worker_id)
        
        return result
        
    except Exception as e:
        debug_print(f"extract_cell_values failed: {type(e).__name__}: {e}", worker_id)
        if safe_mode:
            return {}
        else:
            raise RuntimeError(f"儲存格值提取失敗: {e}")

def full_excel_scan_task(file_path: str, safe_mode: bool = False, worker_id: int = 0,
                        include_formulas: bool = True, include_values: bool = True,
                        batch_size: int = 10000) -> Dict[str, Any]:
    """
    完整 Excel 掃描 (子進程版本) - 替代主進程的 dump_excel_cells_with_timeout
    
    Args:
        file_path: Excel 檔案路徑
        safe_mode: 是否使用安全模式
        worker_id: 工作者 ID
        include_formulas: 是否包含公式
        include_values: 是否包含值
        batch_size: 批次大小
        
    Returns:
        完整的 Excel 資料結構
    """
    debug_print(f"full_excel_scan start file={os.path.basename(file_path)}", worker_id)
    debug_print(f"options: formulas={include_formulas}, values={include_values}, batch={batch_size}", worker_id)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.formula import ArrayFormula
        from datetime import datetime
        
        # 載入工作簿 (公式模式)
        wb_formulas = load_workbook(file_path, read_only=True, data_only=False, keep_vba=False)
        
        # 如果需要值，再載入一次 (值模式)
        wb_values = None
        if include_values:
            try:
                wb_values = load_workbook(file_path, read_only=True, data_only=True, keep_vba=False)
            except Exception as e:
                debug_print(f"values workbook load failed: {e}", worker_id)
                if not safe_mode:
                    raise
        
        result = {}
        
        for ws_idx, ws_formulas in enumerate(wb_formulas.worksheets):
            ws_name = ws_formulas.title
            debug_print(f"processing worksheet {ws_idx+1}/{len(wb_formulas.worksheets)}: '{ws_name}'", worker_id)
            
            try:
                # 獲取對應的值工作表
                ws_values = None
                if wb_values:
                    try:
                        ws_values = wb_values[ws_name]
                    except KeyError:
                        debug_print(f"values worksheet '{ws_name}' not found", worker_id)
                
                # 掃描工作表
                ws_data = _scan_worksheet_complete(
                    ws_formulas, ws_values, worker_id, safe_mode,
                    include_formulas, include_values, batch_size
                )
                
                if ws_data:
                    result[ws_name] = ws_data
                    debug_print(f"worksheet '{ws_name}' completed, {len(ws_data)} cells", worker_id)
                
            except Exception as ws_e:
                debug_print(f"worksheet '{ws_name}' failed: {ws_e}", worker_id)
                if not safe_mode:
                    raise
                continue
        
        # 清理
        wb_formulas.close()
        if wb_values:
            wb_values.close()
        
        total_cells = sum(len(ws_data) for ws_data in result.values())
        debug_print(f"full_excel_scan completed, total cells: {total_cells}", worker_id)
        
        return result
        
    except Exception as e:
        debug_print(f"full_excel_scan failed: {type(e).__name__}: {e}", worker_id)
        if safe_mode:
            return {}
        else:
            raise RuntimeError(f"Excel 完整掃描失敗: {e}")

def _scan_worksheet_complete(ws_formulas, ws_values, worker_id: int, safe_mode: bool,
                            include_formulas: bool, include_values: bool, batch_size: int) -> Dict[str, Dict[str, Any]]:
    """掃描單個工作表的完整資料"""
    
    try:
        max_row = getattr(ws_formulas, 'max_row', 0) or 0
        max_col = getattr(ws_formulas, 'max_column', 0) or 0
        
        if max_row == 0 or max_col == 0:
            return {}
        
        result = {}
        
        # 決定是否需要分批處理
        total_cells = max_row * max_col
        use_batching = total_cells > batch_size
        
        if use_batching:
            debug_print(f"large worksheet ({max_row}x{max_col}), using batch processing", worker_id)
            
            batch_rows = max(1, batch_size // max_col)
            for batch_start in range(1, max_row + 1, batch_rows):
                batch_end = min(batch_start + batch_rows - 1, max_row)
                
                try:
                    batch_data = _extract_cell_data_from_range(
                        ws_formulas, ws_values, batch_start, batch_end, 1, max_col,
                        worker_id, safe_mode, include_formulas, include_values
                    )
                    result.update(batch_data)
                    
                except Exception as batch_e:
                    debug_print(f"batch {batch_start}-{batch_end} failed: {batch_e}", worker_id)
                    if not safe_mode:
                        raise
                    continue
        else:
            # 小文件：直接處理
            result = _extract_cell_data_from_range(
                ws_formulas, ws_values, 1, max_row, 1, max_col,
                worker_id, safe_mode, include_formulas, include_values
            )
        
        return result
        
    except Exception as e:
        debug_print(f"worksheet scan failed: {e}", worker_id)
        if not safe_mode:
            raise
        return {}

def _extract_cell_data_from_range(ws_formulas, ws_values, min_row: int, max_row: int, 
                                 min_col: int, max_col: int, worker_id: int, safe_mode: bool,
                                 include_formulas: bool, include_values: bool) -> Dict[str, Dict[str, Any]]:
    """從指定範圍提取儲存格資料"""
    
    result = {}
    
    try:
        # 遍歷公式工作表
        for row in ws_formulas.iter_rows(min_row=min_row, max_row=max_row, 
                                        min_col=min_col, max_col=max_col, values_only=False):
            for cell in row:
                if cell is None:
                    continue
                
                try:
                    addr = getattr(cell, 'coordinate', None)
                    if not addr:
                        continue
                    
                    cell_data = {}
                    
                    # 提取公式
                    if include_formulas:
                        formula_text = None
                        
                        if hasattr(cell, 'formula') and cell.formula:
                            formula_text = cell.formula
                        elif getattr(cell, 'data_type', None) == 'f':
                            val = getattr(cell, 'value', None)
                            if val is not None:
                                try:
                                    from openpyxl.worksheet.formula import ArrayFormula
                                    if isinstance(val, ArrayFormula):
                                        formula_text = val.text if hasattr(val, 'text') else str(val)
                                    else:
                                        formula_text = str(val)
                                except Exception:
                                    formula_text = str(val)
                        
                        cell_data['formula'] = formula_text
                    
                    # 提取值
                    if include_values:
                        value = None
                        
                        # 優先從值工作表獲取
                        if ws_values:
                            try:
                                value_cell = ws_values[addr]
                                value = value_cell.value
                            except Exception:
                                value = None
                        
                        # 如果值工作表沒有，從公式工作表獲取
                        if value is None:
                            try:
                                value = cell.value
                            except Exception:
                                value = None
                        
                        # 序列化值
                        if value is None:
                            serialized_value = None
                        elif isinstance(value, (int, float, bool, str)):
                            serialized_value = value
                        elif isinstance(value, datetime):
                            serialized_value = value.isoformat()
                        else:
                            serialized_value = str(value)
                        
                        cell_data['value'] = serialized_value
                    
                    # 只有在有資料的情況下才加入結果
                    if any(v is not None for v in cell_data.values()):
                        result[addr] = cell_data
                        
                except Exception as cell_e:
                    if not safe_mode:
                        debug_print(f"cell {cell} processing failed: {cell_e}", worker_id)
                    continue
                    
    except Exception as range_e:
        debug_print(f"range processing failed: {range_e}", worker_id)
        if not safe_mode:
            raise
    
    return result